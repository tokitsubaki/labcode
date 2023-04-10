import openpyxl

# ワークブックの作成
wb = openpyxl.Workbook()
wb.save("test_excel1.xlsx")

# ワークブックの読み込み
wb = openpyxl.load_workbook("test_excel1.xlsx")

# シート名の表示
print(wb.sheetnames)

# 操作対象のシートを選択
ws = wb["Sheet"]
ws = wb.worksheets[0]

# シートを新規作成
wb.create_sheet(title="Sheet2")
print(wb.sheetnames)

wb.create_sheet(title="Sheet3", index=1)
print(wb.sheetnames)

# シートをコピー
wb.copy_worksheet(wb["Sheet"])
print(wb.sheetnames)

# シートを削除
wb.remove(wb["Sheet Copy"])
print(wb.sheetnames)

wb.remove(wb.worksheets[2])
print(wb.sheetnames)

wb.remove(wb.worksheets[-1])
print(wb.sheetnames)

# セルへの書き込み
a1 = ws["A1"]
a1.value = "A1データ"

ws.cell(row=1, column=2, value="B1データ")

# セルの読み取り
a1 = ws["A1"]
a1_data = a1.value
print(a1_data)

b1 = ws.cell(row=1, column=2)
b1_data = b1.value
print(b1_data)

# 範囲指定
a2b4 = ws["A2:C4"]
for rows in a2b4:
    for cell in rows:
        cell.value = cell.column_letter + str(cell.row)
        print(cell.value)

for row in ws.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    print(row_data)

font = openpyxl.styles.Font(name="メイリオ", size=14)
ws["A1"].font = font
ws["B1"].font = font

border = openpyxl.styles.Border(
    top=openpyxl.styles.Side(style="thin", color="000000"),
    right=openpyxl.styles.Side(style="thin", color="000000"),
    bottom=openpyxl.styles.Side(style="thin", color="000000"),
    left=openpyxl.styles.Side(style="thin", color="000000")
)

for row in ws["A2:C4"]:
    for cell in row:
        cell.border = border

fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFFF00")
for row in ws["A2:A4"]:
    for cell in row:
        cell.fill = fill

wb.save("test_excel1.xlsx")




